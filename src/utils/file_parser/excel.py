"""Excel and CSV parser implementation."""

import time
import io
from datetime import datetime
from pathlib import Path
from typing import Dict, List

from .models import Block, ParseResult


async def parse_excel(file_path: str, options: Dict = None) -> ParseResult:
    """Parse Excel file (XLSX).
    
    Args:
        file_path: Path to XLSX file
        options: Parser options
        
    Returns:
        ParseResult
    """
    start_time = time.time()
    options = options or {}
    max_rows = options.get("max_rows", 10000)
    
    file_id = Path(file_path).stem.split("_")[0]
    filename = Path(file_path).name
    
    try:
        import openpyxl
    except ImportError:
        return ParseResult(
            success=False,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            file_id=file_id,
            filename=filename,
            error="openpyxl not installed"
        )
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        blocks = []
        
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            sheet = wb[sheet_name]
            
            # Get all rows (with limit)
            rows = []
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_idx >= max_rows:
                    break
                rows.append([str(cell) if cell is not None else "" for cell in row])
            
            if not rows:
                continue
            
            # Convert to blocks
            table_block = _rows_to_table_block(rows, sheet_idx + 1, sheet_name)
            blocks.append(table_block)
            
            # Also add as paragraphs for small sheets
            if len(rows) <= 10:
                for row_idx, row in enumerate(rows):
                    text = " | ".join(row)
                    blocks.append(Block(
                        chunk_id=f"xlsx_{sheet_idx + 1}_{row_idx + 1}",
                        type="paragraph",
                        content=text,
                        sheet=sheet_name,
                        row_range=f"{row_idx + 1}-{row_idx + 1}",
                        method="openpyxl",
                        confidence=0.95,
                        extracted_at=datetime.now().isoformat()
                    ))
        
        # Generate markdown
        markdown = _blocks_to_markdown(blocks)
        
        return ParseResult(
            success=True,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            file_id=file_id,
            filename=filename,
            markdown=markdown,
            blocks=blocks,
            json={
                "sheets": len(wb.sheetnames),
                "sheet_names": wb.sheetnames
            },
            parse_time_ms=int((time.time() - start_time) * 1000)
        )
        
    except Exception as e:
        return ParseResult(
            success=False,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            file_id=file_id,
            filename=filename,
            error=str(e),
            parse_time_ms=int((time.time() - start_time) * 1000)
        )


async def parse_csv(file_path: str, options: Dict = None) -> ParseResult:
    """Parse CSV file.
    
    Args:
        file_path: Path to CSV file
        options: Parser options
        
    Returns:
        ParseResult
    """
    start_time = time.time()
    options = options or {}
    max_rows = options.get("max_rows", 10000)
    
    file_id = Path(file_path).stem.split("_")[0]
    filename = Path(file_path).name
    
    try:
        import pandas as pd
    except ImportError:
        # Fallback to basic csv
        return await _parse_csv_basic(file_path, options, file_id, filename, start_time)
    
    try:
        # Read CSV with pandas
        df = pd.read_csv(file_path, max_rows=max_rows)
        
        # Convert to rows
        rows = [df.columns.tolist()] + df.values.tolist()
        rows = [[str(cell) for cell in row] for row in rows]
        
        # Create table block
        table_block = _rows_to_table_block(rows, 1, "Sheet1")
        
        # Also add as paragraphs
        blocks = [table_block]
        
        # Add header info
        blocks.append(Block(
            chunk_id="csv_1",
            type="paragraph",
            content=f"Columns: {', '.join(df.columns)}",
            row_range=f"1-{len(df)}",
            method="pandas",
            confidence=0.95,
            extracted_at=datetime.now().isoformat()
        ))
        
        # Add sample rows
        for idx, row in df.head(5).iterrows():
            text = " | ".join(str(v) for v in row.values)
            blocks.append(Block(
                chunk_id=f"csv_{idx + 2}",
                type="paragraph",
                content=text,
                row_range=f"{idx + 2}-{idx + 2}",
                method="pandas",
                confidence=0.95,
                extracted_at=datetime.now().isoformat()
            ))
        
        markdown = _blocks_to_markdown(blocks)
        
        return ParseResult(
            success=True,
            content_type="text/csv",
            file_id=file_id,
            filename=filename,
            markdown=markdown,
            blocks=blocks,
            json={
                "rows": len(df),
                "columns": len(df.columns),
                "column_names": df.columns.tolist()
            },
            parse_time_ms=int((time.time() - start_time) * 1000)
        )
        
    except Exception as e:
        return await _parse_csv_basic(file_path, options, file_id, filename, start_time)


async def _parse_csv_basic(file_path: str, options: Dict, file_id: str, filename: str, start_time: float) -> ParseResult:
    """Basic CSV parsing without pandas."""
    import csv
    
    max_rows = options.get("max_rows", 10000)
    
    try:
        rows = []
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for idx, row in enumerate(reader):
                if idx >= max_rows:
                    break
                rows.append(row)
        
        if not rows:
            return ParseResult(
                success=False,
                content_type="text/csv",
                file_id=file_id,
                filename=filename,
                error="Empty CSV"
            )
        
        table_block = _rows_to_table_block(rows, 1, "Sheet1")
        
        markdown = _blocks_to_markdown([table_block])
        
        return ParseResult(
            success=True,
            content_type="text/csv",
            file_id=file_id,
            filename=filename,
            markdown=markdown,
            blocks=[table_block],
            json={"rows": len(rows)},
            parse_time_ms=int((time.time() - start_time) * 1000)
        )
        
    except Exception as e:
        return ParseResult(
            success=False,
            content_type="text/csv",
            file_id=file_id,
            filename=filename,
            error=str(e),
            parse_time_ms=int((time.time() - start_time) * 1000)
        )


def _rows_to_table_block(rows: List[List[str]], sheet_idx: int, sheet_name: str) -> Block:
    """Convert rows to a table block."""
    if not rows:
        return Block(
            chunk_id=f"xlsx_{sheet_idx}_1",
            type="table",
            content="",
            sheet=sheet_name,
            method="openpyxl",
            confidence=0.95,
            extracted_at=datetime.now().isoformat()
        )
    
    markdown = _table_to_markdown(rows)
    json_table = _table_to_json(rows)
    row_range = f"1-{len(rows)}"
    
    return Block(
        chunk_id=f"xlsx_{sheet_idx}_1",
        type="table",
        content="",
        markdown=markdown,
        table_json=json_table,
        sheet=sheet_name,
        row_range=row_range,
        method="openpyxl",
        confidence=0.95,
        extracted_at=datetime.now().isoformat()
    )


def _table_to_markdown(rows: List[List[str]]) -> str:
    """Convert rows to markdown table."""
    if not rows:
        return ""
    
    # Handle unequal row lengths
    max_cols = max(len(row) for row in rows)
    col_widths = [max(len(str(row[i])) if i < len(row) else 0 for row in rows) for i in range(max_cols)]
    
    lines = []
    
    # Header
    header = "| " + " | ".join(
        str(row[i]).ljust(col_widths[i]) if i < len(row) else " " * col_widths[i]
        for i in range(len(col_widths))
    ) + " |"
    lines.append(header)
    
    # Separator
    sep = "| " + " | ".join("-" * w for w in col_widths) + " |"
    lines.append(sep)
    
    # Data rows
    for row in rows:
        data = "| " + " | ".join(
            str(row[i]).ljust(col_widths[i]) if i < len(row) else " " * col_widths[i]
            for i in range(len(col_widths))
        ) + " |"
        lines.append(data)
    
    return "\n".join(lines)


def _table_to_json(rows: List[List[str]]) -> List[List[str]]:
    """Convert rows to JSON."""
    return [[str(cell) for cell in row] for row in rows]


def _blocks_to_markdown(blocks: List[Block]) -> str:
    """Convert blocks to markdown."""
    md_parts = []
    
    for block in blocks:
        if block.type == "heading":
            level = block.level or 1
            md_parts.append(f"{'#' * level} {block.content}\n")
        elif block.type == "paragraph":
            md_parts.append(f"{block.content}\n")
        elif block.type == "table" and block.markdown:
            md_parts.append(f"\n{block.markdown}\n")
    
    return "\n".join(md_parts)
